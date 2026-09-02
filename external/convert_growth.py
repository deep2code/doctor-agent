#!/usr/bin/env python3
"""Convert growth standard sources into internal/knowledge/data/growth_standards.json.

Sources:
  1. WHO Child Growth Standards z-score tables (xlsx, external/growth/who/)
     weight-for-age (0-5y), length/height-for-age (0-2y + 2-5y),
     head-circumference-for-age (0-5y) — boys & girls.
  2. China WS/T 423-2022 "7岁以下儿童生长标准" (PDF text layer)
     Appendix B SD tables B.1-B.12: weight-for-age, length/height-for-age,
     weight-for-length (0-2y), weight-for-height (2-7y), BMI-for-age (0-7y),
     head-circumference-for-age (0-3y) — boys & girls.
     Plus Table 3 (nutrition assessment rules by SD).

Row schema (numeric arrays always length 7 = -3SD..+3SD):
  {"month": 0, "sd": [2.1, 2.5, 2.9, 3.3, 3.9, 4.4, 5.0]}
  weight-for-length/height rows key on "cm" instead of "month".

Idempotent; run from repo root:
  python3 external/convert_growth.py
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).parent.parent
GROWTH = ROOT / "external" / "growth"
OUT = ROOT / "internal" / "knowledge" / "data" / "growth_standards.json"

SD_KEYS = ["-3sd", "-2sd", "-1sd", "median", "+1sd", "+2sd", "+3sd"]


# ---------------------------------------------------------------- WHO xlsx
SD_COLS = ["SD3neg", "SD2neg", "SD1neg", "SD0", "SD1", "SD2", "SD3"]


def parse_who_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            # wfa sheets: Month,L,M,S,SD3neg..; hcfa/lhfa sheets insert an
            # extra raw-SD column — index by header name, not position.
            if row and row[0] == "Month":
                header = list(row)
            continue
        if row[0] is None:
            continue
        try:
            month = int(row[0])
            vals = [round(float(row[header.index(c)]), 2) for c in SD_COLS]
        except (ValueError, IndexError):
            continue
        rows.append({"month": month, "sd": vals})
    rows.sort(key=lambda r: r["month"])
    return rows


VEL_SD_COLS = ["-3 SD", "-2 SD", "-1 SD", "Median", "1 SD", "2 SD", "3 SD"]
# Intervals: "2-4 mo", "0 – 4 wks", "4 wks – 2 mo" (en/em dashes, mixed units).
RE_INTERVAL = re.compile(r"^(\d+)\s*(wks|mo)?\s*[\u2013\u2014-]\s*(\d+)\s*(wks|mo)$")


def _vel_to_months(n: str, unit: str) -> int:
    """Convert one interval endpoint to whole months (4 wks ≈ 1 mo)."""
    v = int(n)
    if unit == "wks":
        return round(v / 4.345)
    return v


def parse_who_velocity(path: Path) -> list[dict]:
    """Parse one WHO velocity xlsx (rows keyed by interval strings)."""
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and row[0] == "Interval":
                header = list(row)
            continue
        if row[0] is None:
            continue
        m = RE_INTERVAL.match(str(row[0]).strip())
        if not m:
            continue
        try:
            vals = [round(float(row[header.index(c)]), 2) for c in VEL_SD_COLS]
        except (ValueError, IndexError, KeyError):
            continue
        rows.append({"from": _vel_to_months(m.group(1), m.group(2) or "mo"),
                     "to": _vel_to_months(m.group(3), m.group(4) or "mo"),
                     "sd": vals, "interval": str(row[0]).strip()})
    return rows


def build_who_velocity() -> dict:
    """weight (1/2/3/4/6mon), length (2/3/4/6mon), head circumference
    (2/3/4/6mon) — boys & girls, 2009 WHO growth velocity standards.
    Weight tables are in grams per interval; length/head in cm."""
    vel = {"source": "WHO Child Growth Standards: growth velocity based on weight, length and head circumference (2009)",
           "url": "https://www.who.int/tools/child-growth-standards/standards/growth-velocity",
           "weight_unit": "g", "length_head_unit": "cm",
           "note": "按测量间隔(月)选窗口表；体重为区间累计增量(克)，身长/头围为区间累计增量(厘米)"}
    spec = {
        "weight": (["1", "2", "3", "4", "6"], "wv"),
        "length": (["2", "3", "4", "6"], "lv"),
        "head_circumference": (["2", "3", "4", "6"], "hv"),
    }
    for ind, (windows, prefix) in spec.items():
        vel[ind] = {}
        for w in windows:
            entry = {}
            for sex in ("boys", "girls"):
                rows = parse_who_velocity(GROWTH / "who" / "vel" / f"{prefix}_{sex}_{w}.xlsx")
                if rows:
                    entry[sex] = rows
            if entry:
                vel[ind][w] = entry
    return vel


def build_who() -> dict:
    who = {"source": "WHO Child Growth Standards (2006/2007, Multicentre Growth Reference Study)",
           "url": "https://www.who.int/tools/child-growth-standards/standards"}
    who["weight_for_age"] = {"unit": "kg",
        "boys": parse_who_xlsx(GROWTH / "who" / "wfa_boys.xlsx"),
        "girls": parse_who_xlsx(GROWTH / "who" / "wfa_girls.xlsx")}
    # length 0-24 (recumbent) then height 24-60 (standing); both keyed by month
    for sex in ("boys", "girls"):
        rows = []
        for part in ("0_2", "2_5"):
            for r in parse_who_xlsx(GROWTH / "who" / f"lhfa_{sex}_{part}.xlsx"):
                if rows and r["month"] <= rows[-1]["month"]:
                    continue  # 2-5 sheet repeats month 24
                rows.append(r)
        who.setdefault("length_height_for_age", {"unit": "cm", "note": "0-24月为身长(卧位)，24月以上为身高(立位)"})[sex] = rows
    who["head_circumference_for_age"] = {"unit": "cm",
        "note": "WHO 表为 0-60 月",
        "boys": parse_who_xlsx(GROWTH / "who" / "hcfa_boys.xlsx"),
        "girls": parse_who_xlsx(GROWTH / "who" / "hcfa_girls.xlsx")}
    # BMI-for-age 0-60 months (0-2 recumbent + 2-5 standing sheets merged)
    for sex in ("boys", "girls"):
        rows = []
        for part in ("0_2", "2_5"):
            for r in parse_who_xlsx(GROWTH / "who" / "bmi" / f"bmi_{sex}_{part}.xlsx"):
                if rows and r["month"] <= rows[-1]["month"]:
                    continue
                rows.append(r)
        who.setdefault("bmi_for_age", {"unit": "kg/m2", "note": "0-24 月与 24-60 月两表合并"})[sex] = rows
    return who


# ------------------------------------------------------- China WS/T 423 PDF
TABLE_MAP = {
    # table id -> (indicator, sex, key field, unit)
    "B.1": ("weight_for_age", "boys", "month", "kg"),
    "B.2": ("weight_for_age", "girls", "month", "kg"),
    "B.3": ("length_height_for_age", "boys", "month", "cm"),
    "B.4": ("length_height_for_age", "girls", "month", "cm"),
    "B.5": ("weight_for_length", "boys", "cm", "kg"),
    "B.6": ("weight_for_length", "girls", "cm", "kg"),
    "B.7": ("weight_for_height", "boys", "cm", "kg"),
    "B.8": ("weight_for_height", "girls", "cm", "kg"),
    "B.9": ("bmi_for_age", "boys", "month", "kg/m2"),
    "B.10": ("bmi_for_age", "girls", "month", "kg/m2"),
    "B.11": ("head_circumference_for_age", "boys", "month", "cm"),
    "B.12": ("head_circumference_for_age", "girls", "month", "cm"),
}

RE_TABLE_HEAD = re.compile(r"表 (B\.\d+)")
RE_AGE = re.compile(r"^(\d+)\s*岁(?:\s*(\d+)\s*月)?$|^(\d+)\s*月$")
RE_CM = re.compile(r"^(\d{2,3})\s")
RE_NUMS = re.compile(r"^-?\d+(?:\.\d+)?$")


def age_to_month(s: str):
    m = RE_AGE.match(s.strip())
    if not m:
        return None
    if m.group(3) is not None:  # "N 月"
        return int(m.group(3))
    months = int(m.group(1)) * 12
    if m.group(2):
        months += int(m.group(2))
    return months


def parse_china_pdf() -> tuple[dict, list[dict]]:
    r = PdfReader(GROWTH / "wst423-2022.pdf")
    full = []
    for p in r.pages:
        full.append(p.extract_text() or "")

    tables: dict[str, dict[str, list]] = {ind: {} for ind, *_ in TABLE_MAP.values()}
    current = None  # (indicator, sex, key)

    # Table 3 nutrition rules live on pages before appendix B; capture verbatim rows.
    rules = []
    in_rules = False

    for text in full:
        # split into lines but table ids and data may share a line
        for line in text.split("\n"):
            hm = RE_TABLE_HEAD.search(line)
            if hm and hm.group(1) in TABLE_MAP:
                ind, sex, keyf, _unit = TABLE_MAP[hm.group(1)]
                current = (ind, sex, keyf)
                continue
            if "表 3" in line or "标准差评价方法" in line:
                in_rules = True
                current = None
                continue
            if current is None:
                continue

            ind, sex, keyf = current
            # Data line forms (age spans 1-3 tokens, cm is a bare integer):
            #   "9 月 6.9 7.6 ..." | "1 岁 7.4 ..." | "1 岁 1 月 7.5 ..." | "93 11.1 11.8 ..."
            stripped = line.strip()
            key = None
            vals = None
            if "岁" in stripped or "月" in stripped:
                m_ = re.match(
                    r"^(?:(\d+)\s*岁(?:\s*(\d+)\s*月)?|(\d+)\s*月)\s+"
                    r"((?:-?\d+(?:\.\d+)?\s+){6}-?\d+(?:\.\d+)?)\s*$", stripped)
                if not m_:
                    continue
                if m_.group(3) is not None:
                    key = int(m_.group(3))
                else:
                    key = int(m_.group(1)) * 12 + (int(m_.group(2) or 0))
                vals = [float(x) for x in m_.group(4).split()]
            else:
                m_ = re.match(
                    r"^(\d{2,3})\s+((?:-?\d+(?:\.\d+)?\s+){6}-?\d+(?:\.\d+)?)\s*$", stripped)
                if not m_:
                    continue
                key = int(m_.group(1))
                vals = [float(x) for x in m_.group(2).split()]
            if len(vals) != 7:
                continue
            row = {"sd": [round(v, 2) for v in vals]}
            row["month" if keyf == "month" else "cm"] = key
            tables[ind].setdefault(sex, []).append(row)

        if in_rules and "体格测量方法" in text:
            in_rules = False

    # dedupe + sort each series
    for ind in tables:
        for sex in tables[ind]:
            seen = {}
            for row in tables[ind][sex]:
                k = row.get("month", row.get("cm"))
                seen[k] = row
            tables[ind][sex] = sorted(seen.values(), key=lambda r: r.get("month", r.get("cm")))

    # Table 3 rules (hardcoded from PDF page 5 — text layer verified verbatim)
    rules = [
        {"indicator": "年龄别体重", "range": "<-3 SD", "verdict": "重度低体重"},
        {"indicator": "年龄别体重", "range": "-3 SD≤ • <-2 SD", "verdict": "低体重"},
        {"indicator": "年龄别身长/身高", "range": "<-3 SD", "verdict": "重度生长迟缓"},
        {"indicator": "年龄别身长/身高", "range": "-3 SD≤ • <-2 SD", "verdict": "生长迟缓"},
        {"indicator": "身长/身高别体重", "range": "<-3 SD", "verdict": "重度消瘦"},
        {"indicator": "身长/身高别体重", "range": "-3 SD≤ • <-2 SD", "verdict": "消瘦"},
        {"indicator": "年龄别 BMI", "range": "<-3 SD", "verdict": "重度消瘦"},
        {"indicator": "年龄别 BMI", "range": "-3 SD≤ • <-2 SD", "verdict": "消瘦"},
        {"indicator": "年龄别 BMI", "range": "+1 SD≤ • <+2 SD", "verdict": "超重"},
        {"indicator": "年龄别 BMI", "range": "+2 SD≤ • <+3 SD", "verdict": "肥胖"},
        {"indicator": "年龄别 BMI", "range": "≥+3 SD", "verdict": "重度肥胖"},
        {"indicator": "身长/身高别体重", "range": "+1 SD≤ • <+2 SD", "verdict": "超重"},
        {"indicator": "身长/身高别体重", "range": "+2 SD≤ • <+3 SD", "verdict": "肥胖"},
        {"indicator": "身长/身高别体重", "range": "≥+3 SD", "verdict": "重度肥胖"},
        {"indicator": "年龄别体重/身长(身高)/BMI", "range": "-2 SD≤ • <+2 SD", "verdict": "正常"},
    ]
    return tables, rules


# WS/T 456-2014 附录A: 年龄别身高筛查生长迟缓界值(cm)，半岁档 6.0-18.0。
# 来源文字流已与 WS/T 423 的 2-5 岁 -2SD 表互验（男 2 岁 82.0 / 女 80.8 一致）。
SCHOOL_STUNTING_BOYS = {
    "6.0": 106.3, "6.5": 109.5, "7.0": 111.3, "7.5": 112.8, "8.0": 115.4,
    "8.5": 117.6, "9.0": 120.6, "9.5": 123.0, "10.0": 125.2, "10.5": 127.0,
    "11.0": 129.1, "11.5": 130.8, "12.0": 133.1, "12.5": 134.9, "13.0": 136.9,
    "13.5": 138.6, "14.0": 141.9, "14.5": 144.7, "15.0": 149.6, "15.5": 153.6,
    "16.0": 155.1, "16.5": 156.4, "17.0": 156.8, "18.0": 157.1,
}
SCHOOL_STUNTING_GIRLS = {
    "6.0": 105.7, "6.5": 108.0, "7.0": 110.2, "7.5": 111.8, "8.0": 114.5,
    "8.5": 116.8, "9.0": 119.5, "9.5": 121.7, "10.0": 123.9, "10.5": 125.7,
    "11.0": 128.6, "11.5": 131.0, "12.0": 133.6, "12.5": 135.7, "13.0": 138.8,
    "13.5": 141.4, "14.0": 142.9, "14.5": 144.1, "15.0": 145.4, "15.5": 146.5,
    "16.0": 146.8, "16.5": 147.0, "17.0": 147.3, "18.0": 147.5,
}
# 6-17 岁整岁合并筛查表（消瘦界值来自 WS/T 456-2014 附录B 整岁化；
# 超重/肥胖界值与 WS/T 586-2018 一致的体检通用版）。
# band: 消瘦 <=wasting_max < 正常 <=normal_max < 超重 <=overweight_max < 肥胖
SCHOOL_BMI_BOYS = {
    "6": (13.4, 16.7, 18.4), "7": (13.9, 17.3, 19.1), "8": (14.0, 18.0, 20.2),
    "9": (14.1, 18.8, 21.3), "10": (14.4, 19.5, 22.4), "11": (14.9, 20.2, 23.5),
    "12": (15.4, 20.9, 24.6), "13": (15.9, 21.8, 25.6), "14": (16.4, 22.5, 26.3),
    "15": (16.9, 23.0, 26.8), "16": (17.3, 23.4, 27.3), "17": (17.7, 23.7, 27.7),
}
SCHOOL_BMI_GIRLS = {
    "6": (13.1, 16.9, 19.1), "7": (13.4, 17.1, 18.8), "8": (13.6, 18.0, 19.8),
    "9": (13.8, 18.9, 20.9), "10": (14.0, 19.9, 22.0), "11": (14.3, 21.0, 23.2),
    "12": (14.7, 21.8, 24.4), "13": (15.3, 22.5, 25.5), "14": (16.0, 22.9, 26.2),
    "15": (16.6, 23.3, 26.8), "16": (17.0, 23.6, 27.3), "17": (17.2, 23.7, 27.7),
}


def build_school_age() -> dict:
    def bands(src):
        return {a: {"wasting_max": w, "normal_max": n, "overweight_max": o}
                for a, (w, n, o) in src.items()}
    return {
        "source": "WS/T 456-2014《学龄儿童青少年营养不良筛查》+ WS/T 586-2018《学龄儿童青少年超重与肥胖筛查》",
        "urls": [
            "https://www.wiki8.cn/WS.2FT+456.E2.80.942014+xuelingertongqingshaonianyingyangbuliangshaicha_136244/",
            "http://wst.hainan.gov.cn/sjkzx/info/1012/2259.htm",
        ],
        "age_range": "6-18 岁",
        "stunting_height_cm": {
            "note": "身高≤界值判生长迟缓（长期性营养不良）；半岁档",
            "boys": SCHOOL_STUNTING_BOYS, "girls": SCHOOL_STUNTING_GIRLS,
        },
        "bmi_bands": {
            "note": "整岁档合并筛查表：消瘦≤wasting_max<正常≤normal_max<超重≤overweight_max<肥胖；586 官方为半岁档，本表为体检通用整岁版",
            "boys": bands(SCHOOL_BMI_BOYS), "girls": bands(SCHOOL_BMI_GIRLS),
        },
    }


def main():
    who = build_who()
    china_tables, rules = parse_china_pdf()

    china = {
        "source": "WS/T 423-2022《7岁以下儿童生长标准》(国家卫生健康委员会 2022-09-19 发布, 2023-03-01 实施)",
        "url": "https://wsbz.nhc.gov.cn/wsbzw/article/StandardLibrary/2c9081906bf472bd016bf4c1bc700004/2022/11/40900281835df2bd01843764bc0d0019.html",
        "note": "附录B 标准差数值(0-84月)；头围 0-36 月；身长别体重 0-2 岁、身高别体重 2-7 岁按整数 cm",
        "indicators": {},
    }
    units = {v[0]: v[3] for v in TABLE_MAP.values()}
    keyfs = {v[0]: v[2] for v in TABLE_MAP.values()}
    notes = {
        "length_height_for_age": "2 岁以下适用于身长，2～7 岁以下适用于身高",
        "weight_for_length": "0～2 岁以下，按身长(cm)整数",
        "weight_for_height": "2～7 岁以下，按身高(cm)整数",
        "head_circumference_for_age": "0～3 岁",
    }
    for ind, sexes in china_tables.items():
        china["indicators"][ind] = {
            "unit": units[ind],
            "key": "month" if keyfs[ind] == "month" else "cm",
            **({"note": notes[ind]} if ind in notes else {}),
            **sexes,
        }

    doc = {
        "version": "1.1.0",
        "updated": "2026-09-02",
        "who": who,
        "who_velocity": build_who_velocity(),
        "china": china,
        "school_age": build_school_age(),
        "assessment_rules_sd": rules,
        "who_zscore_interpretation": [
            {"range": ">+3 SD", "verdict_en": "obese (BMI/weight-for-height)"},
            {"range": "+2 SD to +3 SD", "verdict_en": "overweight"},
            {"range": "-2 SD to +2 SD", "verdict_en": "normal"},
            {"range": "-3 SD to -2 SD", "verdict_en": "moderate undernutrition"},
            {"range": "<-3 SD", "verdict_en": "severe undernutrition"},
        ],
    }

    OUT.write_text(json.dumps(doc, ensure_ascii=False))
    # stats
    print(f"WHO: wfa boys={len(who['weight_for_age']['boys'])} girls={len(who['weight_for_age']['girls'])}, "
          f"lhfa boys={len(who['length_height_for_age']['boys'])} girls={len(who['length_height_for_age']['girls'])}, "
          f"hcfa boys={len(who['head_circumference_for_age']['boys'])} girls={len(who['head_circumference_for_age']['girls'])}, "
          f"bmi boys={len(who['bmi_for_age']['boys'])} girls={len(who['bmi_for_age']['girls'])}")
    for ind in ("weight", "length", "head_circumference"):
        wins = {w: {s_: len(r) for s_, r in d.items()} for w, d in doc["who_velocity"][ind].items()}
        print(f"WHO velocity {ind}: {wins}")
    print(f"school_age: stunting boys={len(SCHOOL_STUNTING_BOYS)} girls={len(SCHOOL_STUNTING_GIRLS)}, "
          f"bmi bands boys={len(SCHOOL_BMI_BOYS)} girls={len(SCHOOL_BMI_GIRLS)}")
    for ind in china["indicators"]:
        c = china["indicators"][ind]
        print(f"CN {ind}: boys={len(c.get('boys', []))} girls={len(c.get('girls', []))} key={c['key']}")
    print(f"size: {OUT.stat().st_size/1024:.1f} KB -> {OUT}")


if __name__ == "__main__":
    main()
